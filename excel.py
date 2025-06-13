from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
import io

_border_default = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


def _header_sample_stud_group(ws):
    ws['B1'] = 'ТАБЕЛЬ ПОСЕЩАЕМОСТИ ЗАНЯТИЙ'
    ws['B1'].font = Font(bold=True)
    ws['B3'] = 'Дисциплина'
    ws['B3'].font = Font(bold=True)
    for i in range(0, 20):
        ws.cell(column=4 + i, row=3).border = Border(bottom=Side(style='medium'))
    ws['B5'] = 'Преподаватель'
    ws['B5'].font = Font(bold=True)
    for i in range(0, 20):
        ws.cell(column=4 + i, row=5).border = Border(bottom=Side(style='medium'))


def _first_sample_stud_group(ws, group_num, specialty_name):
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['D'].width = 10
    _header_sample_stud_group(ws)
    ws['A8'] = '№'
    ws['B8'] = 'П/Г'
    ws['C8'] = 'ФИО'
    ws['D8'] = '№ билета'
    ws['B7'] = group_num
    ws['B7'].font = Font(bold=True)
    ws['D7'] = specialty_name


def _second_sample_stud_group(ws, group_num, specialty_name):
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['C'].width = 10
    _header_sample_stud_group(ws)
    ws['A8'] = '№'
    ws['B8'] = 'ФИО'
    ws['C8'] = '№ билета'
    ws['B7'] = group_num
    ws['B7'].font = Font(bold=True)
    ws['C7'] = specialty_name


def _create_data_stud_group(group, stud_name_format, sub_group_split):
    def create_data(students, stud_name_format):
        data = [[i + 1 for i in range(len(students))],
                [getattr(s, stud_name_format) for s in
                 students], [s.id for s in students]]

        if group.sub_count > 1:
            data.insert(1, [s.stud_group_subnum for s in students])

        return data

    students = sorted(group.students, key=lambda s: (s.stud_group_subnum if sub_group_split else 0, s.person.surname, s.person.firstname, s.person.middlename))
    if not sub_group_split:
        data = create_data(students, stud_name_format)
    else:
        cur_data = []
        for student in students:
            num = student.stud_group_subnum
            if num == 0:
                num += 1
            if num > len(cur_data):
                while not (len(cur_data) == num):
                    cur_data.append([])
            cur_data[num - 1].append(student)
        data = create_data(sum(cur_data, []), stud_name_format)
    return data


def _fill_worksheet_stud_group(group, ws, stud_name_format, sub_group_split):
    data_group = _create_data_stud_group(group, stud_name_format, sub_group_split)
    if len(data_group) > 3:
        _first_sample_stud_group(ws, f'{group.course} Курс {group.num} Группа', group.specialty.full_name)
        ws.column_dimensions["C"].width = max(len(str(x)) for x in data_group[2]) + 1 if len(data_group[2]) else 20
    else:
        _second_sample_stud_group(ws, f'{group.course} Курс {group.num} Группа', group.specialty.full_name)
        ws.column_dimensions["B"].width = max(len(str(x)) for x in data_group[1]) + 1 if len(data_group[1]) else 20
    start_col = 1

    count_student_cells = 20

    for sub_group in data_group:
        start_row = 9
        for i in range(1, count_student_cells + len(data_group) + 1):
            ws.cell(column=i, row=start_row - 1).border = _border_default
        for student in sub_group:
            ws.cell(column=start_col, row=start_row).value = student
            for i in range(1, count_student_cells+len(data_group)+1):
                ws.cell(column=i, row=start_row).border = _border_default
            start_row += 1
        start_col += 1

    for i in range(0, count_student_cells):
        ws.column_dimensions[get_column_letter(len(data_group)+i+1)].width = 7

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    # A4
    ws.page_setup.paperSize = 9
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3


def create_excel_stud_groups(groups, stud_name_format, sub_group_split):
    wb = Workbook()
    file = io.BytesIO()
    ws = wb.active
    for g in groups:
        ws_n = wb.create_sheet(title=f'{g.course} К{" маг." if g.specialty.education_level == "master" else ""} {g.num} Гр')
        _fill_worksheet_stud_group(g, ws_n, stud_name_format, sub_group_split)
    wb.remove(ws)
    wb.save(file)
    return file


def _fill_line_certificate_of_study(ws, certificate, i):
    ws.cell(i, 1, certificate.print_time.strftime("%d.%m.%Y"))
    ws.cell(i, 2, certificate.print_num)
    ws.cell(i, 3, certificate.specialty.faculty.name_short)

    if certificate.middlename:
        student_name = "%s %s %s" % (certificate.surname, certificate.firstname, certificate.middlename)
    else:
        student_name = "%s %s" % (certificate.surname, certificate.firstname)

    ws.cell(i, 4, student_name)
    course_str_parts = []
    if certificate.course is not None:
        course_str_parts.append(str(certificate.course))

    if certificate.specialty.education_level == 'master':
        course_str_parts.append('маг.')

    if certificate.student_status == 'expelled':
       course_str_parts.append('отчислен' if certificate.student.person.gender == 'M' else 'отчислена')

    if certificate.student_status == 'alumnus':
       course_str_parts.append('выпускник' if certificate.student.person.gender == 'M' else 'выпускница')

    ws.cell(i, 5, " ".join(course_str_parts))
    for j in range(1, 6):
        ws.cell(i, j).border = _border_default


def create_excel_certificates_of_study(certificates):
    wb = Workbook()
    file = io.BytesIO()
    ws = wb.active

    for col in ('A', 'E'):
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['C'].width = 22

    ws['A1']='Дата'
    ws['B1']='Номер'
    ws['C1']='Факультет'
    ws['D1']='Фамилия имя отчество'
    ws['E1']='Курс'

    for cell in ('A1', 'B1', 'C1', 'D1', 'E1'):
        ws[cell].font = Font(bold=True)
        ws[cell].border = _border_default

    i = 1
    prev = None
    last_num = None
    for certificate in certificates:
        if prev is not None and prev.student_id == certificate.student_id:
            ws.cell(i, 2, "%d-%d-%d" % (certificate.specialty.faculty.ID_DEFAULT, last_num, certificate.num))
        else:
            i += 1
            last_num = certificate.num
            _fill_line_certificate_of_study(ws, certificate, i)
        prev = certificate

    for row in ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=0, max_col=2):
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal='left',
            )
    ws.freeze_panes = 'A2'
    wb.save(file)

    return file